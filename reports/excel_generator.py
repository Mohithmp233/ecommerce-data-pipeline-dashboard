"""
excel_generator.py
==================
Generate a weekly executive Excel report from the SQLite warehouse.

Run from the project root:
    python reports/excel_generator.py
"""
from __future__ import annotations

import sqlite3
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from config import REPORTS_DIR, SQLITE_PATH


BRAND_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
BAD_FILL = PatternFill("solid", fgColor="F4CCCC")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def _fetch_rows(conn: sqlite3.Connection, query: str) -> list[dict[str, Any]]:
    """Run a SQL query and return rows as dictionaries."""
    conn.row_factory = sqlite3.Row
    cur = conn.execute(query)
    return [dict(row) for row in cur.fetchall()]


def _write_table(ws, start_row: int, start_col: int, rows: list[dict[str, Any]]) -> int:
    """Write dictionary rows to a worksheet and return the next empty row."""
    if not rows:
        ws.cell(start_row, start_col, "No data available")
        return start_row + 1

    headers = list(rows[0].keys())
    for offset, header in enumerate(headers):
        cell = ws.cell(start_row, start_col + offset, header)
        cell.fill = HEADER_FILL
        cell.font = BOLD_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start_row + 1):
        for col_idx, header in enumerate(headers, start_col):
            cell = ws.cell(row_idx, col_idx, row[header])
            cell.border = THIN_BORDER
            if isinstance(row[header], float):
                cell.number_format = '#,##0.00'
            elif isinstance(row[header], int):
                cell.number_format = '#,##0'
    return start_row + len(rows) + 1


def _format_sheet(ws) -> None:
    """Apply practical widths, freeze panes and alignment to a worksheet."""
    ws.freeze_panes = "A2"
    for column_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
            cell.alignment = Alignment(vertical="top", wrap_text=False)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 32)


def _style_title(ws, title: str, subtitle: str | None = None) -> None:
    """Create a consistent title band at the top of each sheet."""
    ws["A1"] = title
    ws["A1"].fill = BRAND_FILL
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="left")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(italic=True, color="666666")


def _get_kpi_rows(conn: sqlite3.Connection) -> dict[str, Any]:
    """Return headline KPI values for the executive summary sheet."""
    query = """
    WITH order_kpis AS (
        SELECT
            ROUND(SUM(line_total), 2) AS total_revenue,
            COUNT(DISTINCT invoice_no) AS total_orders,
            COUNT(DISTINCT customer_id) AS active_customers,
            ROUND(SUM(line_total) / NULLIF(COUNT(DISTINCT invoice_no), 0), 2) AS avg_order_value,
            ROUND(AVG(data_quality_score), 2) AS data_quality_score
        FROM fact_orders
    ),
    returns AS (
        SELECT COUNT(*) AS return_lines FROM fact_returns
    ),
    top_category AS (
        SELECT p.category
        FROM fact_orders o
        JOIN dim_products p ON o.product_id = p.product_id
        GROUP BY p.category
        ORDER BY SUM(o.line_total) DESC
        LIMIT 1
    ),
    churn AS (
        SELECT COUNT(*) AS churn_risk_count
        FROM (
            SELECT o.customer_id
            FROM fact_orders o
            JOIN dim_date d ON o.date_id = d.date_id
            GROUP BY o.customer_id
            HAVING julianday((SELECT MAX(full_date) FROM dim_date)) - julianday(MAX(d.full_date)) >= 90
        ) x
    ),
    mom AS (
        SELECT
            year,
            month,
            revenue,
            LAG(revenue) OVER (ORDER BY year, month) AS prev_revenue
        FROM (
            SELECT d.year, d.month, SUM(o.line_total) AS revenue
            FROM fact_orders o
            JOIN dim_date d ON o.date_id = d.date_id
            GROUP BY d.year, d.month
        )
    )
    SELECT
        ok.total_revenue,
        ROUND(100.0 * (SELECT revenue - prev_revenue FROM mom WHERE prev_revenue IS NOT NULL ORDER BY year DESC, month DESC LIMIT 1)
              / NULLIF((SELECT prev_revenue FROM mom WHERE prev_revenue IS NOT NULL ORDER BY year DESC, month DESC LIMIT 1), 0), 2) AS mom_growth_pct,
        ok.active_customers,
        ok.avg_order_value,
        ROUND(100.0 * r.return_lines / NULLIF(ok.total_orders, 0), 2) AS return_rate_pct,
        (SELECT category FROM top_category) AS top_category,
        churn.churn_risk_count,
        ok.data_quality_score
    FROM order_kpis ok
    CROSS JOIN returns r
    CROSS JOIN churn;
    """
    rows = _fetch_rows(conn, query)
    return rows[0] if rows else {}


def _add_executive_summary(wb: Workbook, conn: sqlite3.Connection) -> dict[str, Any]:
    """Create the Executive Summary sheet with KPI cards."""
    ws = wb.active
    ws.title = "Executive Summary"
    _style_title(ws, "Executive Summary", f"Generated: {datetime.now():%Y-%m-%d %H:%M}")
    kpis = _get_kpi_rows(conn)

    cards = [
        ("Total Revenue", kpis.get("total_revenue", 0), "$#,##0"),
        ("MoM Growth", (kpis.get("mom_growth_pct") or 0) / 100, "0.0%"),
        ("Active Customers", kpis.get("active_customers", 0), "#,##0"),
        ("Avg Order Value", kpis.get("avg_order_value", 0), "$#,##0.00"),
        ("Return Rate", (kpis.get("return_rate_pct") or 0) / 100, "0.0%"),
        ("Top Category", kpis.get("top_category", "N/A"), "@"),
        ("Churn Risk Count", kpis.get("churn_risk_count", 0), "#,##0"),
        ("Data Quality Score", (kpis.get("data_quality_score") or 0) / 100, "0.0%"),
    ]

    row, col = 4, 1
    for idx, (label, value, number_format) in enumerate(cards):
        r = row + (idx // 4) * 4
        c = col + (idx % 4) * 2
        ws.cell(r, c, label)
        ws.cell(r, c).fill = BRAND_FILL
        ws.cell(r, c).font = WHITE_FONT
        ws.cell(r, c).alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + 1)
        value_cell = ws.cell(r + 1, c, value)
        value_cell.font = Font(bold=True, size=13)
        value_cell.number_format = number_format
        value_cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=r + 1, start_column=c, end_row=r + 2, end_column=c + 1)
        for rr in range(r, r + 3):
            for cc in range(c, c + 2):
                ws.cell(rr, cc).border = THIN_BORDER
                if rr > r:
                    ws.cell(rr, cc).fill = GOOD_FILL if label not in {"Return Rate", "Churn Risk Count"} else WARN_FILL

    _format_sheet(ws)
    return kpis


def _add_revenue_sheet(wb: Workbook, conn: sqlite3.Connection) -> None:
    """Create monthly revenue table, bar chart and below-average highlighting."""
    ws = wb.create_sheet("Revenue Breakdown")
    _style_title(ws, "Revenue Breakdown")
    rows = _fetch_rows(conn, """
        SELECT
            printf('%04d-%02d', d.year, d.month) AS year_month,
            ROUND(SUM(o.line_total), 2) AS revenue,
            COUNT(DISTINCT o.invoice_no) AS orders,
            ROUND(SUM(o.line_total) / NULLIF(COUNT(DISTINCT o.invoice_no), 0), 2) AS avg_order_value
        FROM fact_orders o
        JOIN dim_date d ON o.date_id = d.date_id
        GROUP BY d.year, d.month
        ORDER BY d.year, d.month
    """)
    _write_table(ws, 3, 1, rows)
    if rows:
        avg_revenue = sum(row["revenue"] for row in rows) / len(rows)
        ws.conditional_formatting.add(
            f"B4:B{3 + len(rows)}",
            CellIsRule(operator="lessThan", formula=[str(avg_revenue)], fill=BAD_FILL),
        )
        chart = BarChart()
        chart.title = "Monthly Revenue"
        chart.y_axis.title = "Revenue"
        chart.x_axis.title = "Month"
        data = Reference(ws, min_col=2, min_row=3, max_row=3 + len(rows))
        cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(rows))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 18
        ws.add_chart(chart, "F3")
    _format_sheet(ws)


def _add_customer_segments_sheet(wb: Workbook, conn: sqlite3.Connection) -> None:
    """Create RFM segment counts and revenue table."""
    ws = wb.create_sheet("Customer Segments")
    _style_title(ws, "Customer Segments")
    rows = _fetch_rows(conn, """
        WITH metrics AS (
            SELECT
                o.customer_id,
                CAST(julianday((SELECT MAX(full_date) FROM dim_date)) - julianday(MAX(d.full_date)) AS INTEGER) AS recency_days,
                COUNT(DISTINCT o.invoice_no) AS frequency,
                SUM(o.line_total) AS monetary
            FROM fact_orders o
            JOIN dim_date d ON o.date_id = d.date_id
            GROUP BY o.customer_id
        ),
        scored AS (
            SELECT
                customer_id,
                monetary,
                NTILE(5) OVER (ORDER BY recency_days DESC) AS r_score,
                NTILE(5) OVER (ORDER BY frequency ASC) AS f_score,
                NTILE(5) OVER (ORDER BY monetary ASC) AS m_score
            FROM metrics
        ),
        segments AS (
            SELECT
                customer_id,
                monetary,
                CASE
                    WHEN r_score = 5 AND f_score >= 4 AND m_score >= 4 THEN 'Champions'
                    WHEN f_score >= 4 AND m_score >= 3 THEN 'Loyal'
                    WHEN r_score = 5 AND f_score <= 2 THEN 'New'
                    WHEN r_score <= 2 AND (f_score >= 3 OR m_score >= 3) THEN 'At Risk'
                    WHEN r_score <= 2 THEN 'Lost'
                    ELSE 'Average'
                END AS rfm_segment
            FROM scored
        )
        SELECT
            rfm_segment,
            COUNT(*) AS customers,
            ROUND(SUM(monetary), 2) AS revenue,
            ROUND(AVG(monetary), 2) AS avg_customer_value
        FROM segments
        GROUP BY rfm_segment
        ORDER BY revenue DESC
    """)
    _write_table(ws, 3, 1, rows)
    _format_sheet(ws)


def _add_product_sheet(wb: Workbook, conn: sqlite3.Connection) -> None:
    """Create top product table with margin conditional formatting."""
    ws = wb.create_sheet("Product Performance")
    _style_title(ws, "Product Performance")
    rows = _fetch_rows(conn, """
        SELECT
            p.stock_code,
            p.description,
            p.category,
            ROUND(SUM(o.line_total), 2) AS revenue,
            SUM(o.quantity) AS units_sold,
            ROUND(AVG(p.profit_margin), 2) AS profit_margin_pct
        FROM fact_orders o
        JOIN dim_products p ON o.product_id = p.product_id
        GROUP BY p.stock_code, p.description, p.category
        ORDER BY revenue DESC
        LIMIT 20
    """)
    _write_table(ws, 3, 1, rows)
    if rows:
        margin_col = 6
        ws.conditional_formatting.add(
            f"{get_column_letter(margin_col)}4:{get_column_letter(margin_col)}{3 + len(rows)}",
            CellIsRule(operator="greaterThan", formula=["40"], fill=GOOD_FILL),
        )
        ws.conditional_formatting.add(
            f"{get_column_letter(margin_col)}4:{get_column_letter(margin_col)}{3 + len(rows)}",
            CellIsRule(operator="lessThan", formula=["20"], fill=BAD_FILL),
        )
    _format_sheet(ws)


def _add_alerts_sheet(wb: Workbook, conn: sqlite3.Connection) -> None:
    """Create alert tables for churn, returns, pricing and low stock."""
    ws = wb.create_sheet("Alerts & Flags")
    _style_title(ws, "Alerts & Flags")
    sections = [
        ("Churn Risk Customers", """
            SELECT c.customer_id, c.name, c.email, c.city_tier, c.loyalty_tier,
                   CAST(julianday((SELECT MAX(full_date) FROM dim_date)) - julianday(MAX(d.full_date)) AS INTEGER) AS days_inactive,
                   ROUND(SUM(o.line_total), 2) AS lifetime_revenue
            FROM fact_orders o
            JOIN dim_customers c ON o.customer_id = c.customer_id
            JOIN dim_date d ON o.date_id = d.date_id
            GROUP BY c.customer_id
            HAVING days_inactive >= 90
            ORDER BY lifetime_revenue DESC
            LIMIT 20
        """),
        ("Return Fraud Flags", """
            SELECT r.customer_id, c.name, COUNT(*) AS return_lines,
                   ROUND(SUM(ABS(r.quantity) * r.unit_price), 2) AS returned_value
            FROM fact_returns r
            JOIN dim_customers c ON r.customer_id = c.customer_id
            GROUP BY r.customer_id, c.name
            HAVING return_lines >= 3
            ORDER BY returned_value DESC
            LIMIT 20
        """),
        ("Dark Pattern Products", """
            SELECT p.stock_code, p.description, p.category,
                   ROUND(MAX(o.unit_price) - MIN(o.unit_price), 2) AS price_range,
                   COUNT(DISTINCT o.unit_price) AS price_points
            FROM fact_orders o
            JOIN dim_products p ON o.product_id = p.product_id
            GROUP BY p.stock_code, p.description, p.category
            HAVING price_points > 3
            ORDER BY price_range DESC
            LIMIT 20
        """),
        ("Low Stock Products", """
            SELECT stock_code, description, category, reorder_level, low_stock_flag
            FROM dim_products
            WHERE low_stock_flag = 1
            ORDER BY reorder_level DESC
            LIMIT 20
        """),
    ]
    row = 3
    for title, query in sections:
        ws.cell(row, 1, title)
        ws.cell(row, 1).font = Font(bold=True, size=12)
        ws.cell(row, 1).fill = WARN_FILL
        row = _write_table(ws, row + 1, 1, _fetch_rows(conn, query)) + 2
    _format_sheet(ws)


def generate_weekly_report(db_conn: sqlite3.Connection | str | Path | None = None,
                           output_path: str | Path | None = None) -> tuple[Path, dict[str, Any]]:
    """Generate the weekly Excel report workbook.

    Parameters
    ----------
    db_conn : sqlite3.Connection | str | Path | None
        Existing SQLite connection or database path. Defaults to the
        project SQLite warehouse.
    output_path : str | Path | None
        Optional workbook path. Defaults to reports/weekly_report_YYYY-MM-DD.xlsx.

    Returns
    -------
    tuple[Path, dict[str, Any]]
        Workbook path and a small summary dictionary.
    """
    close_conn = False
    if isinstance(db_conn, sqlite3.Connection):
        conn = db_conn
    else:
        db_path = Path(db_conn) if db_conn is not None else SQLITE_PATH
        conn = sqlite3.connect(db_path)
        close_conn = True

    try:
        REPORTS_DIR.mkdir(parents=True, exist_ok=True)
        out = Path(output_path) if output_path else REPORTS_DIR / f"weekly_report_{datetime.now():%Y-%m-%d}.xlsx"
        wb = Workbook()
        kpis = _add_executive_summary(wb, conn)
        _add_revenue_sheet(wb, conn)
        _add_customer_segments_sheet(wb, conn)
        _add_product_sheet(wb, conn)
        _add_alerts_sheet(wb, conn)
        wb.save(out)
        summary = {
            "output_path": str(out),
            "total_revenue": kpis.get("total_revenue"),
            "active_customers": kpis.get("active_customers"),
            "top_category": kpis.get("top_category"),
            "generated_at": datetime.now().isoformat(timespec="seconds"),
        }
        return out, summary
    finally:
        if close_conn:
            conn.close()


def main() -> None:
    """CLI entry point for report generation."""
    path, summary = generate_weekly_report()
    print(f"Weekly report generated: {path}")
    print(summary)


if __name__ == "__main__":
    main()
